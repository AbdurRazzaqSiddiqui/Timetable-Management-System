from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, reverse, redirect
from django.http import HttpResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from .models import User, University, Department, Section, Subject, Teacher, Timing, Batch, Timetable_components, Timing, Venue
from django.contrib.auth.models import Group
from .email_verification import compose_email,send_email
from functools import wraps
from django.http import HttpResponseForbidden
from openpyxl import Workbook
from datetime import time
from django.core.serializers import serialize
import json
import pickle
import random
from django.db import IntegrityError

def require_verification():
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            # Check if the attribute is true in the request session
            if request.user.verified:
                # Attribute is true, allow access to the view function
                return view_func(request, *args, **kwargs)
            else:
                # Attribute is not true, return forbidden response
                return render(request, 'TMS/verify_email.html',{
                    'user_email':request.session['email']
                }
                )
        return wrapper
    return decorator

# Create your views here.
def login_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('TMS:index'))
    if request.method == 'POST':
        request.session['username'] = request.POST['username']
        request.session['password'] = request.POST['password']
        user = authenticate(request, username=request.session['username'], password=request.session['password'])
        if user is not None:
            login(request,user)
            if not User.objects.get(username=request.session['username']).verified:
                return render(request, 'TMS/verify_email.html',{
                    'user_email':User.objects.get(username=request.session['username']).email
                }
                )
            return HttpResponseRedirect(reverse('TMS:index'))
        else:
            print("Invalid Username or Password")
            return render(request, 'TMS/login_register.html')
    else:
        return render(request, 'TMS/login_register.html')

def register_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect(reverse('TMS:index'))
    if request.method == 'POST':
        request.session['username'] = request.POST['username']
        if University.objects.get(university_name='FAST NUCES').university_email in request.POST['email']:
            request.session['email'] = request.POST['email']
        else:
            print("Please Enter your University Email Address.")
            return render(request, 'TMS/login_register.html')
        request.session['password'] = request.POST['password']
        request.session['confirmation'] = request.POST['confirmation']
        if request.POST['password'] == request.POST['confirmation'] and not (User.objects.filter(email=request.POST['email']).exists()) and not (User.objects.filter(username=request.POST['username']).exists()):
            verification_code = compose_email(request.session['username'],request.session['email'])
            user = User.objects.create_user(request.session['username'],request.session['email'],request.session['password'],is_active=True,is_staff=False, verification_code=verification_code)
            user.save()
            login(request,user)
            if request.POST['user'] == 'student':
                group = Group.objects.get(name='Students')
            elif request.POST['user'] == 'teacher':
                group = Group.objects.get(name='Teachers')
            group.user_set.add(user)
            return render(request, 'TMS/verify_email.html',{
                'user_email':request.session['email']
            })
        else:
            print("Username or Email already taken.")
            return render(request, 'TMS/login_register.html')
    else:
        return render(request, 'TMS/login_register.html')

# def verify_first(request):
#     request.session['staff'] = User.objects.get(username=request.user.username).verified
#     print(request.session['staff'])
#     if not bool(request.session['staff']):
#         print("Hello")
#         request.session['email'] = User.objects.get(username=request.user.username).email
#         return False
#     else:
#         return True

@login_required(login_url='TMS:login')
@require_verification()
def University_Form(request):
    if request.method == 'POST':
        University.objects.create(university_name=request.POST['universityName'],university_email=request.POST['email'],university_location=request.POST['address'],semester_type=request.POST['Semester'])
        uni = University.objects.filter(university_name=request.POST['universityName'])[0]
        for i in range(1,int(request.POST['numDepartments'])+1):
            name1 = f"depName{i}"
            name2 = f"depCode{i}"
            sections = []
            secs = 4
            for j in range(1,secs+1):
                sections.append(f"depName{i}numSections{j}")
            Department.objects.create(university=uni,department_name=request.POST[name1],department_code=request.POST[name2],sections=[int(request.POST[section]) for section in sections])
        return render(request,"TMS/semester_info_form.html")
    else:
        return render(request, "TMS/university_info_form.html")

@login_required(login_url='TMS:login')
@require_verification()
def Semester_Form(request):
    numSubjects = 1
    if request.method == 'POST':
        university = University.objects.all()[0]
        departments = Department.objects.all()
        # if university.semester_type == 'Fall':
        #     batches = [(datetime.date.today().year)-i for i in range(4)]
        # elif university.semester_type == 'Spring':
        #     batches = [(datetime.date.today().year)-1-i for i in range(4)]
        batches = Batch.objects.all()

        for i in range(len(departments)):
            for j in range(len(batches)):
                for k in range(numSubjects):
                    print(f"{i}-{j}")
                    subject_name = f"department{i}-batch{j}-subject{k}"
                    print(request.POST[subject_name])
                    subject_code = f"{subject_name}Code"
                    print(request.POST[subject_code])
                    subject_teacher = f"{subject_name}Teacher"
                    print(request.POST[subject_teacher])
                    print()
                    try:
                        teacher = Teacher.objects.get(teacher_name=request.POST[subject_teacher])
                    except Teacher.DoesNotExist:
                        teacher = Teacher.objects.create(teacher_name=request.POST[subject_teacher])
                    Subject.objects.create(batch=batches[j],subject_name=request.POST[subject_name],subject_code=request.POST[subject_code],subject_teacher=teacher)
        return index(request)
    else:
        return render(request, 'TMS/semester_info_form.html',{
            "departments":[dep.department_name for dep in Department.objects.all()],
            "numDepartments":len(Department.objects.all()),
            "numBatches":len(Batch.objects.all()),
            "numSubjects":numSubjects
        })

@login_required(login_url='TMS:login')
@require_verification()
def index(request):
    # if not verify_first(request):
    #     return render(request, 'TMS/verify_email.html',{
    #         'user_email':request.session['email']
    #     }
    #     )

    opt = True
    if opt:
        # Create_Timetable_Components()
        timetable = Generate_Timetable()
        Save_Timetable(timetable)
        # Generate_Excel_Sheet(timetable)
    else:
        timetable = Read_Timetable()

    # Create_Timings_and_Venues()
    # print(Timetable_components.objects.all())

    return render(request,"TMS/index.html",{
        "timetable":timetable,
        "numVenues":len(Venue.objects.all())
    })

def teacher_timetable(request):
    dep = Department.objects.get(department_code="CS")
    bat = Batch.objects.get(department=dep,year=2024)
    sub = Subject.objects.get(batch=bat,subject_teacher=Teacher.objects.get(teacher_name="a"))
    # print(teacher)
    timetable = Read_Timetable()
    results = Create_Table()
    
    for day in timetable:
        for i in range(len(timetable[day])):
            for j in range(len(timetable[day][i])):
                if timetable[day][i][j] != 0:
                    if timetable[day][i][j].teacher == Teacher.objects.get(teacher_name="a"):
                        results[day][i][j] = timetable[day][i][j]
    
    return render(request,"TMS/teacher_timetable.html",{
        "timetable":results
    })

def verify_email(request):
    user = request.user
    request.session['email'] = User.objects.get(username=user.username).email
    if request.method == 'POST':
        print(user.verification_code)
        print(request.POST['verification_code'])
        if int(request.POST['verification_code']) == int(user.verification_code):
            user.verified = True
            user.save()
            return HttpResponseRedirect(reverse('TMS:index'))
    else:
        return render(request, 'TMS/verify_email.html',{
            'user_email':request.session['email']
        }
        )

def Create_Timetable_Components():
    batches = Batch.objects.all()
    for batch in batches:
        sections = Section.objects.filter(batch=batch)
        subjects = Subject.objects.filter(batch=batch)
        for section in sections:
            for subject in subjects:
                teacher = Teacher.objects.filter(teacher_name=subject.subject_teacher)[0]
                for i in range(subject.credit_hrs):
                    comp = Timetable_components.objects.create(section=section,subject=subject,teacher=teacher)
                # print(comp)

def Create_Timings_and_Venues():
    for i in range(1,21):
        Venue.objects.create(venue_code=f"E{i}",capacity=50)

    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    for i in range(1):
        for j in range(8):
            s_time = time(j+8,0)
            e_time = time(j+8,55)
            Timing.objects.create(day_of_week=days[i],start_time=s_time,end_time=e_time)

def Create_Teachers():
    batches = Batch.objects.all()
    for batch in batches:
        sections = Section.objects.filter(batch=batch)
        subjects = Subject.objects.filter(batch=batch)
        for section in subjects:
            teacher = Teacher.objects.create(teacher_name="Sample")
            teacher.save()
            teacher.subjects.add(Subject.objects.get(batch=batch.pk))
            teacher.save()

def Create_Table():
    timetable = {
        'Monday':[],
        'Tuesday':[],
        'Wednesday':[],
        'Thursday':[],
        'Friday':[]
    }

    for day in timetable:
        for _ in range(len(Venue.objects.all())):
            timetable[day].append([])

    for day in timetable:
        for i in range(len(Venue.objects.all())):
            for _ in range(len(Timing.objects.all())):
                timetable[day][i].append(0)

    return timetable

def Check_Conflict(table, day, j, obj):
    # count = 0
    # for i in range(len(table)):
    #     if table[day][i][j] != 0:
    #         if obj.section == table[day][i][j].section:
    #             count += 1

    for i in range(len(table[day])):
        if table[day][i][j] != 0:
            if obj.section == table[day][i][j].section or obj.teacher == table[day][i][j].teacher:
                return False

    return True

def shuffle_model_instances(queryset):
    primary_keys = list(queryset.values_list('pk', flat=True))
    random.shuffle(primary_keys)
    shuffled_instances = [Timetable_components.objects.get(pk=i) for i in primary_keys]
    return shuffled_instances

def Generate_Timetable():
    table = Create_Table()
    # timetable_objects = Timetable_components.objects.all()
    # timetable_objects = list(timetable_objects)
    # random.shuffle(timetable_objects)
    timetable_objects = shuffle_model_instances(Timetable_components.objects.all())

    flag = False
    for obj in timetable_objects:
        for day in table:
            for i in range(len(table[day])):
                for j in range(len(table[day][i])):
                    if(Check_Conflict(table,day,j,obj) and table[day][i][j] == 0):
                        table[day][i][j] = obj
                        flag = True
                        break
                if(flag):
                    break
            if(flag):
                flag = False
                break
    
    return table

def Print_Table(table):
    for day, data in table.items():
        print(day)
        for row in data:
            for col in row:
                if col == 0: 
                    print("None")
                else:
                    print(f"{col.section}, {col.subject}")
            print()

def Save_Timetable(timetable, filename="data.pkl"):
    # serialized_data = {day: serialize('json', data) for day, data in timetable.items()}
    # with open(filename, 'w') as file:
    #     json.dump(serialized_data, file)
    with open(filename, 'wb') as file:
        pickle.dump(timetable, file)

def Read_Timetable(filename="data.pkl"):
    # with open(filename, 'r') as file:
    #     loaded_data = json.load(file)
    # deserialized_data = {day: list(serialize('json', data)) for day, data in loaded_data.items()}
    # return deserialized_data
    with open(filename, 'rb') as file:
        dictionary = pickle.load(file)
    return dictionary

def Generate_Excel_Sheet(tables):
    wb = Workbook()

    # Iterate over each day and its corresponding table
    for day, table in tables.items():
        ws = wb.create_sheet(title=day)  # Create a new sheet for the day

        # Populate venue names in the first column
        venues = Venue.objects.all()
        for i, venue in enumerate(venues, start=2):
            venue_name = f"{venue.venue_code}"
            tuple_name = f'A{i}'
            ws[tuple_name] = venue_name

        # Populate timing names in the first row
        timings = Timing.objects.all()
        for i, timing in enumerate(timings, start=2):
            timing_name = f"{timing.start_time} - {timing.end_time}"
            tuple_name = f'{chr(64 + i)}1'
            ws[tuple_name] = timing_name

        # Populate table data
        for i in range(len(table)):
            row = f"{i+2}"
            for j in range(len(table[i])):
                column = f"{chr(66+j)}"
                box = column + row
                if table[i][j] == 0:
                    ws[box] = "-"
                else:
                    ws[box] = f"{table[i][j]}"

    wb.save('example.xlsx')

