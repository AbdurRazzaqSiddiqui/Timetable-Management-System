from openpyxl import Workbook

# Create a new workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

# Add data to cells
text = f'BCS-6A\nSir Farrukh Hassan\nSE'
ws['A1'] = "BCS-6A\nSir Farrukh Hassan\nSE"
ws['B1'] = 'BCS-6A\nSir Farrukh Hassan\nSE'
ws['A2'] = 'BCS-6A\nSir Farrukh Hassan\nSE'
ws['B2'] = text
ws['A3'] = 'BCS-6A\nSir Farrukh Hassan\nSE'
ws['B3'] = 'BCS-6A\nSir Farrukh Hassan\nSE'

# Save the workbook
wb.save('example.xlsx')
